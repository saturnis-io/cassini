"""Excel export service for SPC characteristic data.

Builds a 5-sheet openpyxl Workbook from pre-fetched data.
Pure transformation module — no database access.

Sheets:
  1. Measurements — one row per sample, wide format with formulas
  2. Summary Statistics — formula-based stats referencing Measurements
  3. Control Limits — static configuration values
  4. Violations — one row per violation event
  5. Annotations — one row per annotation
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from cassini.db.models.annotation import Annotation
    from cassini.db.models.characteristic import Characteristic
    from cassini.db.models.sample import Sample
    from cassini.db.models.violation import Violation

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

NELSON_RULE_DESCRIPTIONS: dict[int, str] = {
    1: "One point beyond 3\u03c3",
    2: "Nine points same side of center",
    3: "Six points trending",
    4: "Fourteen points alternating",
    5: "Two of three beyond 2\u03c3",
    6: "Four of five beyond 1\u03c3",
    7: "Fifteen points within 1\u03c3",
    8: "Eight points beyond 1\u03c3 both sides",
}

# Styling constants
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_META_FONT = Font(bold=True, color="2F5496")

_VIOLATION_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Row layout for Measurements sheet
_META_ROW_1 = 1  # Characteristic name
_META_ROW_2 = 2  # Hierarchy path
_META_ROW_3 = 3  # Data window description
_BLANK_ROW = 4
_HEADER_ROW = 5
_DATA_START_ROW = 6


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_export_workbook(
    characteristic: Characteristic,
    samples: list[Sample],
    violations_by_sample: dict[int, list[Violation]],
    annotations: list[Annotation],
    hierarchy_path: str,
    data_window_description: str,
) -> io.BytesIO:
    """Build an Excel workbook with 5 sheets from pre-fetched SPC data.

    Returns the workbook serialised to a BytesIO buffer ready for streaming.
    """
    wb = Workbook()

    n = characteristic.subgroup_size or 1
    dp = characteristic.decimal_precision if characteristic.decimal_precision is not None else 3
    num_fmt = f"0.{'0' * dp}" if dp > 0 else "0"

    _build_measurements_sheet(
        wb, characteristic, samples, violations_by_sample, annotations,
        hierarchy_path, data_window_description, n, num_fmt,
    )
    _build_summary_sheet(wb, characteristic, samples, n, num_fmt)
    _build_control_limits_sheet(wb, characteristic, num_fmt)
    _build_violations_sheet(wb, violations_by_sample, samples)
    _build_annotations_sheet(wb, annotations)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Sheet 1 — Measurements
# ---------------------------------------------------------------------------


def _build_measurements_sheet(
    wb: Workbook,
    char: Characteristic,
    samples: list[Sample],
    violations_by_sample: dict[int, list[Violation]],
    annotations: list[Annotation],
    hierarchy_path: str,
    data_window_description: str,
    n: int,
    num_fmt: str,
) -> None:
    ws = wb.active
    ws.title = "Measurements"  # type: ignore[assignment]

    # ---- metadata rows ----
    ws.cell(row=_META_ROW_1, column=1, value=f"Characteristic: {char.name}").font = _META_FONT
    ws.cell(row=_META_ROW_2, column=1, value=f"Location: {hierarchy_path}").font = _META_FONT
    ws.cell(row=_META_ROW_3, column=1, value=f"Data Window: {data_window_description}").font = _META_FONT

    # ---- build column layout ----
    # Fixed columns before measurement values
    fixed_before = ["Sample #", "Timestamp", "Operator", "Batch"]
    # Measurement columns (Meas 1 .. Meas n)
    meas_cols = [f"Meas {i + 1}" for i in range(n)]
    # Computed columns after measurements
    computed_cols = ["Mean", "Range", "Min", "Max"]
    if n > 1:
        computed_cols.append("StdDev")
    # Trailing info columns
    trailing_cols = ["Violations", "Annotations", "Excluded"]

    all_headers = fixed_before + meas_cols + computed_cols + trailing_cols

    # Write header row
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    # Freeze panes below header
    ws.freeze_panes = f"A{_DATA_START_ROW}"

    # Column index helpers (1-based)
    meas_start_col = len(fixed_before) + 1  # first measurement column
    meas_end_col = meas_start_col + n - 1
    mean_col = meas_end_col + 1
    range_col = mean_col + 1
    min_col = range_col + 1
    max_col = min_col + 1
    stddev_col = max_col + 1 if n > 1 else None
    violations_col = (stddev_col + 1) if stddev_col else (max_col + 1)
    annotations_col = violations_col + 1
    excluded_col = annotations_col + 1

    # Build annotation lookup by sample_id for quick access.
    # Period annotations (start_sample_id/end_sample_id) are excluded from
    # per-row display — they span many rows and appear on the Annotations sheet.
    annotation_map: dict[int, list[str]] = {}
    for ann in annotations:
        if ann.sample_id is not None:
            annotation_map.setdefault(ann.sample_id, []).append(ann.text)

    # ---- data rows ----
    for row_offset, sample in enumerate(samples):
        row = _DATA_START_ROW + row_offset

        # Fixed columns
        ws.cell(row=row, column=1, value=sample.id)
        ts = _naive_dt(sample.timestamp) if isinstance(sample.timestamp, datetime) else sample.timestamp
        if isinstance(ts, datetime):
            cell_ts = ws.cell(row=row, column=2, value=ts)
            cell_ts.number_format = "YYYY-MM-DD hh:mm:ss"
        else:
            ws.cell(row=row, column=2, value=str(ts) if ts else "")
        ws.cell(row=row, column=3, value=sample.operator_id or "")
        ws.cell(row=row, column=4, value=sample.batch_number or "")

        # Measurement values — sort by id for consistent column ordering
        sorted_measurements = sorted(sample.measurements, key=lambda m: m.id)
        for i in range(n):
            col = meas_start_col + i
            if i < len(sorted_measurements):
                cell_m = ws.cell(row=row, column=col, value=sorted_measurements[i].value)
                cell_m.number_format = num_fmt
            else:
                ws.cell(row=row, column=col, value="")

        # Formula columns — reference measurement cells
        meas_start_letter = get_column_letter(meas_start_col)
        meas_end_letter = get_column_letter(meas_end_col)
        meas_range = f"{meas_start_letter}{row}:{meas_end_letter}{row}"

        # Mean = AVERAGE of measurements
        cell_mean = ws.cell(
            row=row, column=mean_col,
            value=f"=AVERAGE({meas_range})",
        )
        cell_mean.number_format = num_fmt

        # Range = MAX - MIN
        cell_range = ws.cell(
            row=row, column=range_col,
            value=f"=MAX({meas_range})-MIN({meas_range})",
        )
        cell_range.number_format = num_fmt

        # Min
        cell_min = ws.cell(
            row=row, column=min_col,
            value=f"=MIN({meas_range})",
        )
        cell_min.number_format = num_fmt

        # Max
        cell_max = ws.cell(
            row=row, column=max_col,
            value=f"=MAX({meas_range})",
        )
        cell_max.number_format = num_fmt

        # StdDev (only if subgroup > 1)
        if stddev_col is not None:
            cell_sd = ws.cell(
                row=row, column=stddev_col,
                value=f"=STDEV.S({meas_range})",
            )
            cell_sd.number_format = num_fmt

        # Violations text
        sample_violations = violations_by_sample.get(sample.id, [])
        if sample_violations:
            viol_text = "; ".join(
                NELSON_RULE_DESCRIPTIONS.get(v.rule_id, v.rule_name or f"Rule {v.rule_id}")
                for v in sample_violations
            )
            ws.cell(row=row, column=violations_col, value=viol_text)
        else:
            ws.cell(row=row, column=violations_col, value="")

        # Annotations text
        sample_annotations = annotation_map.get(sample.id, [])
        if sample_annotations:
            ws.cell(row=row, column=annotations_col, value="; ".join(sample_annotations))
        else:
            ws.cell(row=row, column=annotations_col, value="")

        # Excluded flag
        ws.cell(row=row, column=excluded_col, value="Yes" if sample.is_excluded else "")

        # Red fill on violation rows
        if sample_violations:
            for c in range(1, excluded_col + 1):
                ws.cell(row=row, column=c).fill = _VIOLATION_FILL

    # ---- auto-width columns ----
    _auto_width(ws, max_col=excluded_col)


# ---------------------------------------------------------------------------
# Sheet 2 — Summary Statistics
# ---------------------------------------------------------------------------


def _build_summary_sheet(
    wb: Workbook,
    char: Characteristic,
    samples: list[Sample],
    n: int,
    num_fmt: str,
) -> None:
    ws = wb.create_sheet("Summary Statistics")

    # Column layout helpers from Measurements sheet
    fixed_before_count = 4  # Sample #, Timestamp, Operator, Batch
    mean_col = fixed_before_count + n + 1
    range_col = mean_col + 1

    mean_letter = get_column_letter(mean_col)
    range_letter = get_column_letter(range_col)

    # Measurement column letters (for individual value formulas)
    meas_start_col = fixed_before_count + 1
    meas_end_col = meas_start_col + n - 1

    total_data_rows = len(samples)
    last_data_row = _DATA_START_ROW + total_data_rows - 1 if total_data_rows > 0 else _DATA_START_ROW

    # Header
    ws.cell(row=1, column=1, value="Statistic").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=1).alignment = _HEADER_ALIGNMENT
    ws.cell(row=1, column=2, value="Value").font = _HEADER_FONT
    ws.cell(row=1, column=2).fill = _HEADER_FILL
    ws.cell(row=1, column=2).alignment = _HEADER_ALIGNMENT
    ws.cell(row=1, column=3, value="Formula / Source").font = _HEADER_FONT
    ws.cell(row=1, column=3).fill = _HEADER_FILL
    ws.cell(row=1, column=3).alignment = _HEADER_ALIGNMENT

    ws.freeze_panes = "A2"

    row = 2
    data_range_mean = f"Measurements!{mean_letter}{_DATA_START_ROW}:{mean_letter}{last_data_row}"
    data_range_range = f"Measurements!{range_letter}{_DATA_START_ROW}:{range_letter}{last_data_row}"

    def _write_stat(label: str, value: str | float | None, note: str = "") -> int:
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        val_cell = ws.cell(row=row, column=2)
        if isinstance(value, str) and value.startswith("="):
            val_cell.value = value
            val_cell.number_format = num_fmt
        elif isinstance(value, (int, float)):
            val_cell.value = value
            val_cell.number_format = num_fmt
        else:
            val_cell.value = value
        ws.cell(row=row, column=3, value=note)
        current = row
        row += 1
        return current

    # Sample count — COUNTA on the mean column minus header rows
    _write_stat(
        "Sample Count",
        f"=COUNTA({data_range_mean})",
        f"COUNTA of Mean column ({mean_letter})",
    )

    # X-bar (Grand Mean)
    xbar_row = _write_stat(
        "X-bar (Grand Mean)",
        f"=AVERAGE({data_range_mean})",
        f"AVERAGE of Mean column ({mean_letter})",
    )

    # R-bar (Average Range)
    rbar_row = _write_stat(
        "R-bar (Average Range)",
        f"=AVERAGE({data_range_range})",
        f"AVERAGE of Range column ({range_letter})",
    )

    # Within-subgroup sigma (R-bar/d2) — from backend's stored_sigma.
    # This is the correct sigma for Cp/Cpk (within-subgroup capability).
    # For n=1 subgroups, stored_sigma uses moving range / d2.
    within_sigma_row = _write_stat(
        "Within-Subgroup Sigma",
        char.stored_sigma,
        "R-bar / d2 (from SPC engine)",
    )

    # Overall Std Dev (of all individual measurements) — for Pp/Ppk
    if n == 1:
        meas_letter = get_column_letter(meas_start_col)
        overall_range = f"Measurements!{meas_letter}{_DATA_START_ROW}:{meas_letter}{last_data_row}"
    else:
        first_meas_letter = get_column_letter(meas_start_col)
        last_meas_letter = get_column_letter(meas_end_col)
        overall_range = (
            f"Measurements!{first_meas_letter}{_DATA_START_ROW}"
            f":{last_meas_letter}{last_data_row}"
        )
    overall_sd_row = _write_stat(
        "Overall Std Dev",
        f"=STDEV.S({overall_range})",
        "STDEV of all individual measurements",
    )

    # Blank separator
    row += 1

    # Specification limits (static)
    usl_row = _write_stat("USL", char.usl, "Upper Specification Limit")
    lsl_row = _write_stat("LSL", char.lsl, "Lower Specification Limit")
    _write_stat("Target", char.target_value, "Target Value")

    # Blank separator
    row += 1

    # Capability indices — only when both USL and LSL exist
    if char.usl is not None and char.lsl is not None:
        ws.cell(row=row, column=1, value="Process Capability").font = _META_FONT
        row += 1

        xbar_cell = f"B{xbar_row}"
        within_sigma_cell = f"B{within_sigma_row}"
        overall_sd_cell = f"B{overall_sd_row}"
        usl_cell = f"B{usl_row}"
        lsl_cell = f"B{lsl_row}"

        # Cp = (USL - LSL) / (6 * within_sigma)
        _write_stat(
            "Cp",
            f"=({usl_cell}-{lsl_cell})/(6*{within_sigma_cell})",
            "(USL - LSL) / (6 * Within σ)",
        )

        # Cpu = (USL - X-bar) / (3 * within_sigma)
        cpu_row = _write_stat(
            "Cpu",
            f"=({usl_cell}-{xbar_cell})/(3*{within_sigma_cell})",
            "(USL - X̄) / (3 * Within σ)",
        )

        # Cpl = (X-bar - LSL) / (3 * within_sigma)
        cpl_row = _write_stat(
            "Cpl",
            f"=({xbar_cell}-{lsl_cell})/(3*{within_sigma_cell})",
            "(X̄ - LSL) / (3 * Within σ)",
        )

        # Cpk = MIN(Cpu, Cpl)
        _write_stat(
            "Cpk",
            f"=MIN(B{cpu_row},B{cpl_row})",
            "MIN(Cpu, Cpl)",
        )

        # Pp = (USL - LSL) / (6 * overall_stddev)
        _write_stat(
            "Pp",
            f"=({usl_cell}-{lsl_cell})/(6*{overall_sd_cell})",
            "(USL - LSL) / (6 * Overall σ)",
        )

        # Ppu = (USL - X-bar) / (3 * overall_stddev)
        ppu_row = _write_stat(
            "Ppu",
            f"=({usl_cell}-{xbar_cell})/(3*{overall_sd_cell})",
            "(USL - X̄) / (3 * Overall σ)",
        )

        # Ppl = (X-bar - LSL) / (3 * overall_stddev)
        ppl_row = _write_stat(
            "Ppl",
            f"=({xbar_cell}-{lsl_cell})/(3*{overall_sd_cell})",
            "(X̄ - LSL) / (3 * Overall σ)",
        )

        # Ppk = MIN(Ppu, Ppl)
        _write_stat(
            "Ppk",
            f"=MIN(B{ppu_row},B{ppl_row})",
            "MIN(Ppu, Ppl)",
        )

    _auto_width(ws, max_col=3)


# ---------------------------------------------------------------------------
# Sheet 3 — Control Limits
# ---------------------------------------------------------------------------


def _build_control_limits_sheet(
    wb: Workbook,
    char: Characteristic,
    num_fmt: str,
) -> None:
    ws = wb.create_sheet("Control Limits")

    # Header
    for col_idx, header in enumerate(["Parameter", "Value"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    ws.freeze_panes = "A2"

    rows: list[tuple[str, object]] = [
        ("Characteristic", char.name),
        ("Data Type", char.data_type),
        ("Chart Type", char.chart_type or ("X-bar/R" if char.data_type == "variable" else char.attribute_chart_type or "")),
        ("Subgroup Size", char.subgroup_size),
        ("UCL", char.ucl),
        ("Center Line", char.stored_center_line),
        ("LCL", char.lcl),
        ("USL", char.usl),
        ("LSL", char.lsl),
        ("Target Value", char.target_value),
        ("Stored Sigma", char.stored_sigma),
        ("Decimal Precision", char.decimal_precision),
        ("Short-Run Mode", char.short_run_mode or "None"),
        ("Distribution Method", char.distribution_method or "Normal"),
    ]

    for row_idx, (label, value) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            val_cell.number_format = num_fmt

    _auto_width(ws, max_col=2)


# ---------------------------------------------------------------------------
# Sheet 4 — Violations
# ---------------------------------------------------------------------------


def _build_violations_sheet(
    wb: Workbook,
    violations_by_sample: dict[int, list[Violation]],
    samples: list[Sample],
) -> None:
    ws = wb.create_sheet("Violations")

    headers = [
        "Sample #", "Timestamp", "Rule", "Rule Description",
        "Severity", "Acknowledged", "Ack By", "Ack Reason",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    ws.freeze_panes = "A2"

    row = 2
    # Iterate samples in order so violations appear sorted by sample
    for sample in samples:
        sample_violations = violations_by_sample.get(sample.id, [])
        for v in sample_violations:
            ws.cell(row=row, column=1, value=sample.id)
            ts = _naive_dt(sample.timestamp) if isinstance(sample.timestamp, datetime) else sample.timestamp
            if isinstance(ts, datetime):
                ts_cell = ws.cell(row=row, column=2, value=ts)
                ts_cell.number_format = "YYYY-MM-DD hh:mm:ss"
            else:
                ws.cell(row=row, column=2, value=str(ts) if ts else "")
            ws.cell(row=row, column=3, value=f"Rule {v.rule_id}")
            ws.cell(
                row=row, column=4,
                value=NELSON_RULE_DESCRIPTIONS.get(v.rule_id, v.rule_name or f"Rule {v.rule_id}"),
            )
            ws.cell(row=row, column=5, value=v.severity)
            ws.cell(row=row, column=6, value="Yes" if v.acknowledged else "No")
            ws.cell(row=row, column=7, value=v.ack_user or "")
            ws.cell(row=row, column=8, value=v.ack_reason or "")
            row += 1

    _auto_width(ws, max_col=len(headers))


# ---------------------------------------------------------------------------
# Sheet 5 — Annotations
# ---------------------------------------------------------------------------


def _build_annotations_sheet(
    wb: Workbook,
    annotations: list[Annotation],
) -> None:
    ws = wb.create_sheet("Annotations")

    headers = ["Type", "Sample #", "Text", "Author", "Created"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    ws.freeze_panes = "A2"

    for row_idx, ann in enumerate(annotations, start=2):
        ws.cell(row=row_idx, column=1, value=ann.annotation_type)
        ws.cell(row=row_idx, column=2, value=ann.sample_id if ann.sample_id else "")
        ws.cell(row=row_idx, column=3, value=ann.text)
        ws.cell(row=row_idx, column=4, value=ann.created_by or "")
        ts = _naive_dt(ann.created_at) if isinstance(ann.created_at, datetime) else ann.created_at
        if isinstance(ts, datetime):
            ts_cell = ws.cell(row=row_idx, column=5, value=ts)
            ts_cell.number_format = "YYYY-MM-DD hh:mm:ss"
        else:
            ws.cell(row=row_idx, column=5, value=str(ts) if ts else "")

    _auto_width(ws, max_col=len(headers))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _naive_dt(dt: datetime | None) -> datetime | None:
    """Strip timezone info for openpyxl (Excel does not support tz-aware datetimes)."""
    if dt is None:
        return None
    return dt.replace(tzinfo=None) if dt.tzinfo is not None else dt


def _auto_width(ws: object, max_col: int) -> None:
    """Set column widths based on content, with reasonable bounds."""
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:  # type: ignore[index]
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(val))
            except Exception:
                pass
        # Clamp between 8 and 40 characters, add padding
        adjusted = min(max(max_length + 2, 8), 40)
        ws.column_dimensions[col_letter].width = adjusted  # type: ignore[union-attr]
