"""CSV/Excel file import service for SPC data.

Handles file parsing, column detection, type inference, validation,
and mapping for bulk data import into characteristics.
"""

import csv
import io
from datetime import datetime
from typing import Any

# Date formats to try when parsing timestamps (ordered by likelihood)
DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y",
]

MAX_PREVIEW_ROWS = 10
MAX_SAMPLE_VALUES = 5


def parse_file(file_content: bytes, filename: str) -> dict:
    """Parse a CSV or Excel file and return column metadata with preview rows.

    Args:
        file_content: Raw file bytes.
        filename: Original filename (used for extension detection).

    Returns:
        Dict with keys: columns, row_count, preview_rows.

    Raises:
        ValueError: If file format is unsupported or parsing fails.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        rows = _parse_csv(file_content)
    elif ext in ("xlsx", "xls"):
        rows = _parse_excel(file_content)
    else:
        raise ValueError(f"Unsupported file format: .{ext}. Use .csv or .xlsx")

    if not rows:
        raise ValueError("File contains no data")

    # First row is treated as headers
    headers = rows[0]
    data_rows = rows[1:]

    if not data_rows:
        raise ValueError("File contains headers but no data rows")

    # Build column metadata
    columns = []
    for col_idx, header in enumerate(headers):
        col_values = [
            str(row[col_idx]).strip() if col_idx < len(row) else ""
            for row in data_rows
        ]
        # Filter out empty values for type detection
        non_empty = [v for v in col_values if v]
        detected_type = _detect_type(non_empty) if non_empty else "empty"
        sample_values = non_empty[:MAX_SAMPLE_VALUES]

        columns.append({
            "name": str(header).strip() if header else f"Column {col_idx + 1}",
            "index": col_idx,
            "sample_values": sample_values,
            "detected_type": detected_type,
        })

    preview_rows = [
        [str(cell).strip() if cell is not None else "" for cell in row]
        for row in data_rows[:MAX_PREVIEW_ROWS]
    ]

    return {
        "columns": columns,
        "row_count": len(data_rows),
        "preview_rows": preview_rows,
    }


def _parse_csv(content: bytes) -> list[list[str]]:
    """Parse CSV content into a list of rows.

    Handles BOM encoding via utf-8-sig.
    """
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(cell.strip() for cell in row)]


def _parse_excel(content: bytes) -> list[list[Any]]:
    """Parse Excel (.xlsx) content into a list of rows.

    Uses openpyxl in read-only mode for memory efficiency.
    """
    from openpyxl import load_workbook

    wb = load_workbook(
        filename=io.BytesIO(content),
        read_only=True,
        data_only=True,
    )
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no active worksheet")

        rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if any(cell is not None for cell in row):
                rows.append(list(row))
        return rows
    finally:
        wb.close()


def _detect_type(values: list[str]) -> str:
    """Detect the most likely type for a column based on sample values.

    Tries numeric first, then datetime, then falls back to string.
    """
    if not values:
        return "empty"

    # Check numeric
    numeric_count = 0
    for v in values:
        try:
            float(v.replace(",", ""))
            numeric_count += 1
        except (ValueError, AttributeError):
            pass

    if numeric_count / len(values) >= 0.8:
        return "numeric"

    # Check datetime
    datetime_count = 0
    for v in values:
        if _try_parse_timestamp(v) is not None:
            datetime_count += 1

    if datetime_count / len(values) >= 0.8:
        return "datetime"

    return "string"


def _try_parse_timestamp(value: str) -> datetime | None:
    """Try to parse a string as a datetime using common formats.

    Returns the parsed datetime or None if no format matched.
    """
    value = value.strip()
    if not value:
        return None

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _parse_timestamp(value: str) -> datetime:
    """Parse a string as a datetime. Raises ValueError if unparseable."""
    result = _try_parse_timestamp(value)
    if result is None:
        raise ValueError(f"Cannot parse timestamp: '{value}'")
    return result


def validate_and_map(
    parsed: dict,
    column_mapping: dict[str, int],
    data_type: str,
) -> dict:
    """Validate and map parsed file data according to column mapping.

    Args:
        parsed: Output from parse_file().
        column_mapping: Maps target fields to column indices.
            Variable data: {"value": idx} or {"values": [idx1, idx2, ...]}
            Optional: {"timestamp": idx, "batch_number": idx, "operator_id": idx}
            Attribute data: {"defect_count": idx}, optional {"sample_size": idx, "units_inspected": idx}
        data_type: "variable" or "attribute".

    Returns:
        Dict with keys: valid_rows, warnings, error_rows, total_rows, valid_count.
    """
    valid_rows: list[dict] = []
    warnings: list[str] = []
    error_rows: list[dict] = []

    if data_type == "variable":
        valid_rows, warnings, error_rows = _validate_variable(parsed, column_mapping)
    elif data_type == "attribute":
        valid_rows, warnings, error_rows = _validate_attribute(parsed, column_mapping)
    else:
        raise ValueError(f"Unsupported data_type: {data_type}")

    return {
        "valid_rows": valid_rows,
        "warnings": warnings,
        "error_rows": error_rows,
        "total_rows": parsed["row_count"],
        "valid_count": len(valid_rows),
    }


def _validate_variable(
    parsed: dict,
    column_mapping: dict[str, int | list[int]],
) -> tuple[list[dict], list[str], list[dict]]:
    """Validate and map rows for variable data import."""
    valid_rows: list[dict] = []
    warnings: list[str] = []
    error_rows: list[dict] = []

    # Determine value column(s)
    value_cols: list[int]
    if "values" in column_mapping and isinstance(column_mapping["values"], list):
        value_cols = column_mapping["values"]
    elif "value" in column_mapping:
        value_cols = [column_mapping["value"]]
    else:
        raise ValueError("Column mapping must include 'value' or 'values' for variable data")

    timestamp_col = column_mapping.get("timestamp")
    batch_col = column_mapping.get("batch_number")
    operator_col = column_mapping.get("operator_id")

    data_rows = parsed.get("data_rows", parsed.get("preview_rows", []))

    for row_idx, row in enumerate(data_rows):
        row_num = row_idx + 2  # 1-indexed, +1 for header row
        try:
            mapped = _map_variable_row(
                row, row_num, value_cols, timestamp_col, batch_col, operator_col
            )
            valid_rows.append(mapped)
        except ValueError as e:
            error_rows.append({"row": row_num, "error": str(e)})

    if error_rows and len(error_rows) < len(data_rows):
        warnings.append(
            f"{len(error_rows)} of {len(data_rows)} rows had errors and will be skipped"
        )

    return valid_rows, warnings, error_rows


def _map_variable_row(
    row: list[str],
    row_num: int,
    value_cols: list[int],
    timestamp_col: int | None,
    batch_col: int | None,
    operator_col: int | None,
) -> dict:
    """Map a single row to variable sample fields."""
    measurements: list[float] = []
    for col_idx in value_cols:
        if col_idx >= len(row):
            raise ValueError(f"Row {row_num}: missing column {col_idx}")
        raw = str(row[col_idx]).strip().replace(",", "")
        if not raw:
            raise ValueError(f"Row {row_num}: empty measurement value in column {col_idx}")
        try:
            measurements.append(float(raw))
        except ValueError:
            raise ValueError(f"Row {row_num}: non-numeric value '{row[col_idx]}' in column {col_idx}")

    result: dict[str, Any] = {"measurements": measurements}

    if timestamp_col is not None and timestamp_col < len(row):
        ts_raw = str(row[timestamp_col]).strip()
        if ts_raw:
            try:
                result["timestamp"] = _parse_timestamp(ts_raw)
            except ValueError:
                # Warn but don't fail — use server default
                result["timestamp"] = None

    if batch_col is not None and batch_col < len(row):
        val = str(row[batch_col]).strip()
        if val:
            result["batch_number"] = val

    if operator_col is not None and operator_col < len(row):
        val = str(row[operator_col]).strip()
        if val:
            result["operator_id"] = val

    return result


def _validate_attribute(
    parsed: dict,
    column_mapping: dict[str, int],
) -> tuple[list[dict], list[str], list[dict]]:
    """Validate and map rows for attribute data import."""
    valid_rows: list[dict] = []
    warnings: list[str] = []
    error_rows: list[dict] = []

    if "defect_count" not in column_mapping:
        raise ValueError("Column mapping must include 'defect_count' for attribute data")

    defect_col = column_mapping["defect_count"]
    sample_size_col = column_mapping.get("sample_size")
    units_col = column_mapping.get("units_inspected")
    timestamp_col = column_mapping.get("timestamp")
    batch_col = column_mapping.get("batch_number")
    operator_col = column_mapping.get("operator_id")

    data_rows = parsed.get("data_rows", parsed.get("preview_rows", []))

    for row_idx, row in enumerate(data_rows):
        row_num = row_idx + 2
        try:
            mapped = _map_attribute_row(
                row, row_num, defect_col, sample_size_col, units_col,
                timestamp_col, batch_col, operator_col,
            )
            valid_rows.append(mapped)
        except ValueError as e:
            error_rows.append({"row": row_num, "error": str(e)})

    if error_rows and len(error_rows) < len(data_rows):
        warnings.append(
            f"{len(error_rows)} of {len(data_rows)} rows had errors and will be skipped"
        )

    return valid_rows, warnings, error_rows


def _map_attribute_row(
    row: list[str],
    row_num: int,
    defect_col: int,
    sample_size_col: int | None,
    units_col: int | None,
    timestamp_col: int | None,
    batch_col: int | None,
    operator_col: int | None,
) -> dict:
    """Map a single row to attribute sample fields."""
    if defect_col >= len(row):
        raise ValueError(f"Row {row_num}: missing defect_count column")

    raw = str(row[defect_col]).strip().replace(",", "")
    if not raw:
        raise ValueError(f"Row {row_num}: empty defect_count value")
    try:
        defect_count = int(float(raw))
    except ValueError:
        raise ValueError(f"Row {row_num}: non-numeric defect_count '{row[defect_col]}'")

    if defect_count < 0:
        raise ValueError(f"Row {row_num}: negative defect_count")

    result: dict[str, Any] = {"defect_count": defect_count}

    if sample_size_col is not None and sample_size_col < len(row):
        raw = str(row[sample_size_col]).strip().replace(",", "")
        if raw:
            try:
                result["sample_size"] = int(float(raw))
            except ValueError:
                raise ValueError(f"Row {row_num}: non-numeric sample_size")

    if units_col is not None and units_col < len(row):
        raw = str(row[units_col]).strip().replace(",", "")
        if raw:
            try:
                result["units_inspected"] = int(float(raw))
            except ValueError:
                raise ValueError(f"Row {row_num}: non-numeric units_inspected")

    if timestamp_col is not None and timestamp_col < len(row):
        ts_raw = str(row[timestamp_col]).strip()
        if ts_raw:
            try:
                result["timestamp"] = _parse_timestamp(ts_raw)
            except ValueError:
                result["timestamp"] = None

    if batch_col is not None and batch_col < len(row):
        val = str(row[batch_col]).strip()
        if val:
            result["batch_number"] = val

    if operator_col is not None and operator_col < len(row):
        val = str(row[operator_col]).strip()
        if val:
            result["operator_id"] = val

    return result
